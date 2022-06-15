import time
import unittest
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
from Funciones.Funciones import FuncionesGlobales
from Funciones.PageLogin import PaginaLogin

class Funexcel():
    def __init__(self, driver):
        self.driver = driver

    def getRowCount(file, path, sheetName):
        Worbook = openpyxl.load_workbook(path)
        sheet = Worbook[sheetName]
        return (sheet.max_row)

    def getColumnCount(file, sheetName):
        Worbook = openpyxl.load_workbook(file)
        sheet = Worbook[sheetName]
        return (sheet.max_column)

    def readData(file, path, sheetName, rownum, columnno):
        Worbook = openpyxl.load_workbook(path)
        sheet = Worbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    def writeData(file, path, sheetName, rownum, columnno, data):
        Worbook = openpyxl.load_workbook(path)
        sheet = Worbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        Worbook.save(path)